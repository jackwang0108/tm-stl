"""
从新3D数据中制作数据集

    @Time    : 2024/11/12
    @Author  : JackWang
    @File    : align_intensity.py
    @IDE     : VsCode
"""

# Standard Library
from pathlib import Path
from collections import defaultdict
from argparse import ArgumentParser, Namespace

# Third-Party Library
import pandas as pd
from openpyxl import load_workbook

# My Library
from make_cls_datasets import copy_files


def get_args() -> Namespace:

    parser = ArgumentParser(prog="从新3D数据中制作数据集")

    parser.add_argument(
        "-i",
        "--ipath",
        type=str,
        required=True,
        help="源STL数据文件夹",
    )

    parser.add_argument(
        "-o",
        "--opath",
        type=str,
        default="./output",
        help="目标STL数据文件夹",
    )

    parser.add_argument(
        "-x",
        "--xlsx",
        type=str,
        required=True,
        help="强度xlsx表格",
    )

    return parser.parse_args()


def make_xlsx(
    xlsx_path: Path,
    color_mapper: dict[int, tuple[int, str]] = {
        8: (0, "Brock"),
        9: (1, "Dendritic"),
        6: (2, "Flake"),
        2: (3, "Strip"),
    },
) -> pd.DataFrame:
    assert xlsx_path.exists(), f"{xlsx_path} 不存在"

    df = pd.read_excel(xlsx_path, header=0)

    def get_row_colors(xlsx_path: Path) -> tuple[int, int]:

        sheet = load_workbook(xlsx_path).active

        colors = []
        for row in sheet.iter_rows():
            if (cell := row[0]).fill.patternType == "solid":
                color = cell.fill.start_color.index
                colors.append((cell.row, color))

        return colors

    colors = get_row_colors(xlsx_path)
    type_cols = pd.DataFrame.from_dict(
        {
            "名称": [color_mapper[i[1]][1] for i in colors],
            "类别": [color_mapper[i[1]][0] for i in colors],
        }
    )

    df = pd.concat((df, type_cols), axis=1)

    return df


def get_file_mapping(src_dir: Path, dst_dir: Path, types: pd.Series) -> pd.DataFrame:
    assert src_dir.is_dir() and src_dir.exists(), f"{src_dir} 不存在或者不是目录"

    files = sorted(src_dir.glob("*.stl"), key=lambda x: int(x.stem))

    typed_files = defaultdict(list)
    for type, filenames in zip(types, files):
        typed_files[type].append(filenames)

    file_mappings = []
    for type, filelist in typed_files.items():
        file_mappings.extend(
            (src_file, dst_dir / f"{type}/{idx}.stl")
            for idx, src_file in enumerate(filelist)
        )

    return pd.DataFrame(file_mappings, columns=["源文件", "目标文件"])


def main():
    args = get_args()

    assert (ipath := Path(args.ipath)).exists(), f"{ipath} 不存在"

    df = make_xlsx(Path(args.xlsx))
    file_mappings = get_file_mapping(ipath, opath := Path(args.opath), df.iloc[:, -2])
    copy_files(file_mappings)

    pd.concat((file_mappings, df), axis=1).to_excel(opath / "分类-强度数据集.xlsx")


if __name__ == "__main__":
    main()
